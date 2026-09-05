from __future__ import annotations

import os
import re
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
root = Path(tempfile.mkdtemp(prefix="twd-replenishment-report-"))
os.environ["TWD_DATA_DIR"] = str(root)
os.environ["TWD_SESSION_SECRET"] = "replenishment-report-test-secret"

from fastapi.testclient import TestClient  # noqa: E402
from order_system.database import dumps_json  # noqa: E402
from order_system.web.app import app, repo  # noqa: E402
from order_system.web.repository import ORDER_COLUMNS  # noqa: E402


def csrf(html: str) -> str:
    match = re.search(r'name="csrf" value="([^"]+)"', html)
    assert match
    return match.group(1)


def login(client: TestClient, username: str, password: str) -> None:
    page = client.get("/login")
    response = client.post(
        "/login",
        data={"csrf": csrf(page.text), "username": username, "password": password},
        follow_redirects=False,
    )
    assert response.status_code == 303


def logout(client: TestClient, page_html: str) -> None:
    response = client.post("/logout", data={"csrf": csrf(page_html)}, follow_redirects=False)
    assert response.status_code == 303


def order_payload(order_no: str, salesman: str) -> dict[str, object]:
    data = {column: None for column in ORDER_COLUMNS}
    data.update({
        "order_type": "新订单",
        "salesman": salesman,
        "order_no": order_no,
        "product_name": "补数测试产品",
        "order_date": "2026-09-01",
        "delivery_date": "2026-09-10",
        "quantity": 100,
        "spare_quantity": 0,
        "quantity_unit": "个",
        "unit_price": 1,
        "price_tiers_json": dumps_json([]),
        "extra_fee": 0,
        "order_prefix_no": 1,
        "materials_json": dumps_json([]),
        "plating_json": dumps_json([]),
        "accessories_json": dumps_json([]),
        "polishing_json": dumps_json([]),
        "coloring_json": dumps_json([]),
        "resin_json": dumps_json([]),
        "packaging_json": dumps_json([]),
        "image_paths_json": dumps_json([]),
        "component_parts_json": dumps_json([]),
    })
    return data


with TestClient(app) as client:
    repo.create_user("admin", "admin-pass", "admin", display_name="管理员")
    repo.create_user("prod", "prod-pass-1", "production", display_name="生管员")
    repo.create_user("sales-a", "sales-pass", "sales", display_name="业务甲")
    repo.create_user("workshop", "workshop-pass", "workshop", display_name="车间")
    first_id, _ = repo.create_order(order_payload("TWD1-260901801", "业务甲"))
    second_id, _ = repo.create_order(order_payload("TWD1-260901802", "业务乙"))
    production_user = repo.get_user(2)
    admin_user = repo.get_user(1)
    assert production_user and admin_user
    first_request = repo.create_replenishment_request(first_id, production_user, 12, "首批补数")
    second_request = repo.create_replenishment_request(second_id, production_user, 8, "次批补数")
    repo.review_edit_request(first_request, admin_user, True, "同意")
    repo.review_edit_request(second_request, admin_user, True, "同意")
    with repo.connect(write=True) as conn:
        conn.execute("UPDATE order_edit_requests SET reviewed_at = '2026-09-03 04:00:00' WHERE id = ?", (first_request,))
        conn.execute("UPDATE order_edit_requests SET reviewed_at = '2026-09-04 04:00:00' WHERE id = ?", (second_request,))

    approved = repo.list_edit_requests("approved")
    serial_by_id = {item["id"]: item["created_order_no"] for item in approved}
    first_serial = serial_by_id[first_request]
    second_serial = serial_by_id[second_request]

    login(client, "admin", "admin-pass")
    admin_page = client.get("/replenishments")
    assert admin_page.status_code == 200
    assert "补数单" in admin_page.text and '<a href="/replenishments">补数单</a>' in admin_page.text
    for heading in ("流水号", "数量", "补数人", "原因", "时间"):
        assert heading in admin_page.text
    assert first_serial in admin_page.text and second_serial in admin_page.text
    assert "12" in admin_page.text and "8" in admin_page.text and "生管员" in admin_page.text
    assert "首批补数" in admin_page.text and "次批补数" in admin_page.text
    filtered = client.get("/replenishments?reported_from=2026-09-03&reported_to=2026-09-03")
    assert filtered.status_code == 200 and first_serial in filtered.text and second_serial not in filtered.text
    exported = client.post(
        "/replenishments/export",
        data={"csrf": csrf(filtered.text), "reported_from": "2026-09-03", "reported_to": "2026-09-03"},
    )
    assert exported.status_code == 200
    assert "spreadsheetml.sheet" in exported.headers["content-type"]
    workbook_path = root / "replenishments.xlsx"
    workbook_path.write_bytes(exported.content)
    sheet = load_workbook(workbook_path).active
    assert [sheet.cell(1, column).value for column in range(1, 6)] == ["流水号", "数量", "补数人", "原因", "时间"]
    assert sheet.max_row == 2 and sheet.cell(2, 1).value == first_serial
    assert sheet.cell(2, 2).value == 12 and sheet.cell(2, 3).value == "生管员"
    logout(client, admin_page.text)

    login(client, "prod", "prod-pass-1")
    production_page = client.get("/replenishments")
    assert production_page.status_code == 200 and first_serial in production_page.text and second_serial in production_page.text
    logout(client, production_page.text)

    login(client, "sales-a", "sales-pass")
    sales_page = client.get("/replenishments")
    assert sales_page.status_code == 200 and first_serial in sales_page.text and second_serial not in sales_page.text
    logout(client, sales_page.text)

    login(client, "workshop", "workshop-pass")
    assert client.get("/replenishments").status_code == 403

print(f"replenishment report smoke ok: {root}")
