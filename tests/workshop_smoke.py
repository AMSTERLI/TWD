from __future__ import annotations

import os
import re
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
root = Path(tempfile.mkdtemp(prefix="twd-workshop-"))
os.environ["TWD_DATA_DIR"] = str(root)
os.environ["TWD_SESSION_SECRET"] = "workshop-test-secret-long-enough"
os.environ["TWD_WORKSHOP_MOLD_PASSWORD"] = "mold-pass-123"
os.environ["TWD_WORKSHOP_CUTTER_PASSWORD"] = "cutter-pass-123"
os.environ["TWD_WORKSHOP_PRESS_PASSWORD"] = "press-pass-123"
os.environ["TWD_WORKSHOP_CRYSTAL_PASSWORD"] = "crystal-pass-123"
os.environ["TWD_WORKSHOP_PACKAGING_PASSWORD"] = "packaging-pass-123"
os.environ["TWD_WORKSHOP_POLISHING_PASSWORD"] = "polishing-pass-123"
os.environ["TWD_WORKSHOP_PAINTING_PASSWORD"] = "painting-pass-123"
os.environ["TWD_WORKSHOP_DIECAST_PASSWORD"] = "diecast-pass-123"

from fastapi.testclient import TestClient  # noqa: E402
from order_system.database import dumps_json  # noqa: E402
from order_system.web.app import app, repo  # noqa: E402
from order_system.web.repository import ORDER_COLUMNS  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from PIL import Image as PilImage  # noqa: E402
from order_system.web.settings import IMAGES_DIR  # noqa: E402


def csrf(html: str) -> str:
    match = re.search(r'name="csrf" value="([^"]+)"', html)
    assert match
    return match.group(1)


def payload(order_no: str) -> dict[str, object]:
    data = {column: None for column in ORDER_COLUMNS}
    data.update({
        "order_type": "\u65b0\u8ba2\u5355",
        "salesman": "\u6768\u5a1f",
        "order_no": order_no,
        "product_name": "\u6d4b\u8bd5\u5fbd\u7ae0",
        "order_date": "2026-07-21",
        "delivery_date": "2026-07-30",
        "quantity": 100,
        "spare_quantity": 0,
        "quantity_unit": "\u4e2a",
        "order_prefix_no": 1,
        "materials_json": dumps_json([]),
        "plating_json": dumps_json([]),
        "accessories_json": dumps_json([]),
        "polishing_json": dumps_json([]),
        "coloring_json": dumps_json([]),
        "resin_json": dumps_json([]),
        "packaging_json": dumps_json([]),
        "image_paths_json": dumps_json([]),
        "component_parts_json": dumps_json([]),
    })
    return data


def assert_workshop_detail_pdf_only(html: str, order_id: int, hidden_unit_price: str) -> None:
    assert "pdf-preview" in html and f"/orders/{order_id}/pdf" in html
    assert "detail-grid" not in html
    assert "workflow-line" not in html
    assert "craft-grid" not in html
    assert hidden_unit_price not in html
    assert "&#36710;&#38388;&#25253;&#21040;&#35760;&#24405;" not in html
    assert "外发记录" not in html

static_js = Path("order_system/web/static/app.js").read_text(encoding="utf-8")
assert "????????" not in static_js
assert "以下订单已录入，普通单不允许重复录入" in static_js or "\\u4ee5\\u4e0b\\u8ba2\\u5355\\u5df2\\u5f55\\u5165\\uff0c\\u666e\\u901a\\u5355\\u4e0d\\u5141\\u8bb8\\u91cd\\u590d\\u5f55\\u5165" in static_js
assert "确认继续保存吗" in static_js or "\\u786e\\u8ba4\\u7ee7\\u7eed\\u4fdd\\u5b58\\u5417" in static_js



with TestClient(app) as client:
    repo.create_user("admin", "admin-pass-123", "admin")
    repo.create_user("workshop", "workshop-pass-123", "workshop", display_name="\u8f66\u95f4")
    IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    press_thumb_name = "press-thumb.png"
    PilImage.new("RGB", (24, 18), color=(20, 120, 200)).save(IMAGES_DIR / press_thumb_name)
    mold_payload = payload("TWD1-260721101")
    mold_payload["height_mm"] = "1"
    mold_payload["width_mm"] = "20"
    mold_payload["thickness_mm"] = "5"
    order_id, order_no = repo.create_order(mold_payload)
    cutter_order_id, cutter_order_no = repo.create_order(payload("TWD1-260721102"))
    press_payload = payload("TWD1-260721103")
    press_payload["image_paths_json"] = dumps_json([press_thumb_name])
    press_order_id, press_order_no = repo.create_order(press_payload)
    press_sort_order_id, press_sort_order_no = repo.create_order(payload("TWD1-260721099"))
    crystal_order_id, crystal_order_no = repo.create_order(payload("TWD1-260721105"))
    packaging_order_id, packaging_order_no = repo.create_order(payload("TWD1-260721106"))
    polishing_order_id, polishing_order_no = repo.create_order(payload("TWD1-260721107"))
    painting_order_id, painting_order_no = repo.create_order(payload("TWD1-260721108"))
    diecast_order_id, diecast_order_no = repo.create_order(payload("TWD1-260721109"))
    auto_qty_payload = payload("TWD1-260721104")
    auto_qty_payload["spare_quantity"] = 15
    _, auto_qty_order_no = repo.create_order(auto_qty_payload)

    admin_login_page = client.get("/login")
    admin_login = client.post(
        "/login",
        data={"csrf": csrf(admin_login_page.text), "username": "admin", "password": "admin-pass-123"},
        follow_redirects=False,
    )
    assert admin_login.status_code == 303
    admin_home = client.get("/workshop")
    assert admin_home.status_code == 200 and "/workshop/mold/unlock" not in admin_home.text
    admin_mold = client.get("/workshop/mold")
    assert admin_mold.status_code == 200 and "data-workshop-scan" in admin_mold.text
    admin_cutter = client.get("/workshop/cutter")
    assert admin_cutter.status_code == 200 and "data-workshop-scan" in admin_cutter.text
    client.post("/logout", data={"csrf": csrf(admin_mold.text)})

    login_page = client.get("/login")
    login = client.post(
        "/login",
        data={"csrf": csrf(login_page.text), "username": "workshop", "password": "workshop-pass-123"},
        follow_redirects=False,
    )
    assert login.status_code == 303 and login.headers["location"] == "/workshop"
    assert client.get("/orders").status_code == 403

    home = client.get("/workshop")
    assert home.status_code == 200 and "/workshop/mold/unlock" in home.text and "/workshop/cutter/unlock" in home.text
    assert "/workshop/press" in home.text and "/workshop/crystal" in home.text and "/workshop/packaging" in home.text
    assert "/workshop/polishing" in home.text and "/workshop/painting" in home.text and "/workshop/diecast" in home.text
    bad_unlock = client.post(
        "/workshop/mold/unlock",
        data={"csrf": csrf(home.text), "password": "bad"},
        follow_redirects=False,
    )
    assert bad_unlock.status_code == 403
    unlock = client.post(
        "/workshop/mold/unlock",
        data={"csrf": csrf(home.text), "password": "mold-pass-123"},
        follow_redirects=False,
    )
    assert unlock.status_code == 303 and unlock.headers["location"] == "/workshop/mold"

    mold = client.get("/workshop/mold")
    assert mold.status_code == 200 and "data-workshop-scan" in mold.text
    assert "2D+\u80cc\u5b57" in mold.text and "3D+\u80cc\u5b57" in mold.text and ">\u80cc\u5b57<" in mold.text
    report = client.post(
        "/workshop/mold",
        data={"csrf": csrf(mold.text), "order_no": [order_no], "material": ["锌"], "size_text": ["50MM"], "spec": ["2D+\u80cc\u5b57"], "quantity": ["2"], "unit_price": ["10.5"], "record_type": ["normal"]},
        follow_redirects=False,
    )
    assert report.status_code == 303
    records = repo.order_workshop_records(order_id)
    assert len(records) == 1
    assert records[0]["department_name"] == "\u523b\u6a21"
    assert records[0]["quantity"] == 2
    assert records[0]["material"] == "锌"
    assert records[0]["size_text"] == "50MM"
    assert records[0]["spec"] == "2D+\u80cc\u5b57"
    assert records[0]["record_type"] == "normal"
    assert abs(records[0]["unit_price"] - 10.5) < 1e-9
    cutter_unlock = client.post(
        "/workshop/cutter/unlock",
        data={"csrf": csrf(home.text), "password": "cutter-pass-123"},
        follow_redirects=False,
    )
    assert cutter_unlock.status_code == 303 and cutter_unlock.headers["location"] == "/workshop/cutter"
    cutter = client.get("/workshop/cutter")
    assert cutter.status_code == 200 and "data-workshop-scan" in cutter.text
    assert '<select name="note_text" required>' in cutter.text and 'data-note-preset' not in cutter.text
    assert '<option value="\u65e0">\u65e0</option>' in cutter.text
    assert '<option value="\u7279\u6b8a">\u7279\u6b8a</option>' in cutter.text
    assert "\u5185\u52071\u652f" not in cutter.text
    cutter_report = client.post(
        "/workshop/cutter",
        data={"csrf": csrf(cutter.text), "order_no": [cutter_order_no], "size_text": ["35MM"], "note_text": ["\u7279\u6b8a"], "quantity": ["1"], "unit_price": ["8.8"], "record_type": ["normal"]},
        follow_redirects=False,
    )
    assert cutter_report.status_code == 303
    cutter_records = repo.order_workshop_records(cutter_order_id)
    assert len(cutter_records) == 1
    assert cutter_records[0]["department_name"] == "\u5207\u5200"
    assert cutter_records[0]["quantity"] == 1
    assert cutter_records[0]["size_text"] == "35MM"
    assert cutter_records[0]["note_text"] == "\u7279\u6b8a"
    assert cutter_records[0]["record_type"] == "normal"
    assert abs(cutter_records[0]["unit_price"] - 8.8) < 1e-9
    cutter_list = client.get("/workshop/cutter")
    assert cutter_list.status_code == 200 and cutter_order_no in cutter_list.text
    assert "&#23610;&#23544;" in cutter_list.text and "&#22791;&#27880;" in cutter_list.text
    assert "35MM" in cutter_list.text and "\u7279\u6b8a" in cutter_list.text and "&#27491;&#24120;" in cutter_list.text
    assert 'data-delete-url="/workshop/cutter/records/' not in cutter_list.text
    assert 'data-workshop-quantity-url="/workshop/cutter/records/' in cutter_list.text
    assert "data-selected-amount-total" in cutter_list.text and 'data-amount="8.80"' in cutter_list.text
    assert 'data-selection-amount-total-all="8.80"' in cutter_list.text
    press_unlock = client.post(
        "/workshop/press/unlock",
        data={"csrf": csrf(home.text), "password": "press-pass-123"},
        follow_redirects=False,
    )
    assert press_unlock.status_code == 303 and press_unlock.headers["location"] == "/workshop/press"
    press = client.get("/workshop/press")
    assert press.status_code == 200 and "touch-piecework-panel" in press.text
    assert "workshop-press-page" in press.text
    assert "data-touch-keypad" in press.text and "data-touch-number" in press.text
    assert "data-touch-integer" in press.text and "data-touch-scale" in press.text
    assert 'type="date"' not in press.text
    assert 'name="date_range" value="month"' in press.text
    assert 'name="date_range" value="week"' in press.text
    assert 'name="date_range" value="day"' in press.text
    assert 'name="mold_fee"' in press.text and 'list="mold-fee-options-press"' in press.text and '<option value="8"></option>' in press.text
    assert 'employee-button active' not in press.text
    assert 'data-batch-workshop-price' in press.text
    press_employees = ["\u5f90\u5c71\u7acb", "\u5218\u9053\u6797", "\u6881\u8d3b\u6821", "\u79e6\u5e94\u57ce", "\u66fe\u51e4\u5a25", "\u519c\u7231\u67f3"]
    for employee in press_employees:
        assert f'data-employee-value="{employee}"' in press.text
        assert f'<option value="{employee}"' in press.text
    press_report = client.post(
        "/workshop/press",
        data={"csrf": csrf(press.text), "employee_name": ["\u5f90\u5c71\u7acb,\u5218\u9053\u6797"], "order_no": [press_order_no], "quantity": ["120"], "unit_price": ["0.08"], "mold_fee": ["9"]},
        follow_redirects=False,
    )
    assert press_report.status_code == 303
    press_records = repo.order_workshop_records(press_order_id)
    assert len(press_records) == 2
    assert {row["operator_name"] for row in press_records} == {"\u5f90\u5c71\u7acb", "\u5218\u9053\u6797"}
    for row in press_records:
        assert row["department_name"] == "\u51b2\u538b"
        assert row["quantity"] == 60
        assert abs(row["unit_price"] - 0.08) < 1e-9
        assert abs(row["mold_fee"] - 4.5) < 1e-9
        assert abs(row["amount"] - 9.3) < 1e-9
    press_list = client.get("/workshop/press")
    assert press_list.status_code == 200 and ">\u5f90\u5c71\u7acb<" in press_list.text and press_order_no in press_list.text
    assert "data-selected-amount-total" in press_list.text and 'data-amount="9.30"' in press_list.text
    assert "&#37329;&#39069;" in press_list.text and "9.30" in press_list.text
    assert 'data-selection-amount-total-all="18.60"' in press_list.text
    assert 'data-workshop-quantity-url="/workshop/press/records/' not in press_list.text
    press_filtered = client.get("/workshop/press?employee_name=%E5%BE%90%E5%B1%B1%E7%AB%8B")
    assert press_filtered.status_code == 200 and press_order_no in press_filtered.text and ">\u5f90\u5c71\u7acb<" in press_filtered.text
    press_filtered_second = client.get("/workshop/press?employee_name=%E5%88%98%E9%81%93%E6%9E%97")
    assert press_filtered_second.status_code == 200 and press_order_no in press_filtered_second.text
    press_filtered_empty = client.get("/workshop/press?employee_name=%E6%A2%81%E8%B4%BB%E6%A0%A1")
    assert press_filtered_empty.status_code == 200 and press_order_no not in press_filtered_empty.text
    history_total = client.get(f"/workshop/press/history?order_no={press_order_no}")
    assert history_total.status_code == 200 and history_total.json()["record"]["quantity"] == 100
    assert history_total.json()["record"]["existing_workshop_record"] is True
    auto_qty = client.get(f"/workshop/press/history?order_no={auto_qty_order_no}")
    assert auto_qty.status_code == 200
    assert auto_qty.json()["record"]["quantity"] == 115
    assert auto_qty.json()["record"]["unit_price"] == 0
    assert auto_qty.json()["record"]["existing_workshop_record"] is False
    press_reference_payload = payload(press_order_no)
    press_reference_payload.update({"quantity": 40, "spare_quantity": 5, "_manual_order_no": True})
    repo.create_order(press_reference_payload)
    press_old_reference_payload = payload(press_order_no)
    press_old_reference_payload.update({"order_date": "2026-04-01", "quantity": 999, "spare_quantity": 1, "_manual_order_no": True})
    repo.create_order(press_old_reference_payload)
    press_export = client.post(
        "/workshop/press/export",
        data={"csrf": csrf(press_list.text), "selected_ids": [str(row["id"]) for row in press_records]},
    )
    assert press_export.status_code == 200
    press_workbook_path = root / "press-export.xlsx"
    press_workbook_path.write_bytes(press_export.content)
    press_sheet = load_workbook(press_workbook_path).active
    assert press_sheet.cell(row=1, column=5).value == "\u5458\u5de5"
    assert press_sheet.cell(row=1, column=7).value == "\u53c2\u8003\u6570\u91cf"
    assert press_sheet.cell(row=1, column=12).value == "\u4ea7\u54c1\u7f29\u7565\u56fe"
    assert len(getattr(press_sheet, "_images", [])) == 2
    assert press_sheet.cell(row=2, column=1).value == press_order_no
    assert {press_sheet.cell(row=row_index, column=5).value for row_index in (2, 3)} == {"\u5f90\u5c71\u7acb", "\u5218\u9053\u6797"}
    assert press_sheet.cell(row=2, column=6).value == 60
    assert press_sheet.cell(row=3, column=6).value == 60
    assert press_sheet.cell(row=2, column=7).value == 145
    assert press_sheet.cell(row=3, column=7).value == 145
    assert abs(sum(press_sheet.cell(row=row_index, column=9).value for row_index in (2, 3)) - 9) < 1e-9
    assert abs(sum(press_sheet.cell(row=row_index, column=10).value for row_index in (2, 3)) - 18.6) < 1e-9
    press_sort_report = client.post(
        "/workshop/press",
        data={"csrf": csrf(press_list.text), "employee_name": ["\u6881\u8d3b\u6821"], "order_no": [press_sort_order_no], "quantity": ["20"], "unit_price": ["0.1"], "mold_fee": ["0"]},
        follow_redirects=False,
    )
    assert press_sort_report.status_code == 303
    press_sort_records = repo.order_workshop_records(press_sort_order_id)
    press_sort_export = client.post(
        "/workshop/press/export",
        data={"csrf": csrf(press_list.text), "selected_ids": [str(press_records[0]["id"]), str(press_sort_records[0]["id"]), str(press_records[1]["id"])]},
    )
    assert press_sort_export.status_code == 200
    press_sort_workbook_path = root / "press-sort-export.xlsx"
    press_sort_workbook_path.write_bytes(press_sort_export.content)
    press_sort_sheet = load_workbook(press_sort_workbook_path).active
    assert [press_sort_sheet.cell(row=row_index, column=1).value for row_index in (2, 3, 4)] == [press_sort_order_no, press_order_no, press_order_no]
    crystal_unlock = client.post(
        "/workshop/crystal/unlock",
        data={"csrf": csrf(home.text), "password": "crystal-pass-123"},
        follow_redirects=False,
    )
    assert crystal_unlock.status_code == 303 and crystal_unlock.headers["location"] == "/workshop/crystal"
    crystal = client.get("/workshop/crystal")
    assert crystal.status_code == 200 and "data-workshop-scan" in crystal.text
    assert 'name="unit_price"' not in crystal.text and 'data-workshop-employees' in crystal.text
    assert "玻璃" in crystal.text and "磁铁" in crystal.text and "配件" in crystal.text
    crystal_report = client.post(
        "/workshop/crystal",
        data={"csrf": csrf(crystal.text), "employee_name": ["玻璃"], "order_no": [crystal_order_no], "quantity": ["12"]},
        follow_redirects=False,
    )
    assert crystal_report.status_code == 303
    crystal_records = repo.order_workshop_records(crystal_order_id)
    assert len(crystal_records) == 1
    assert crystal_records[0]["department_name"] == "\u6676\u9762"
    assert crystal_records[0]["quantity"] == 12
    assert abs(crystal_records[0]["unit_price"] - 0) < 1e-9
    crystal_list = client.get("/workshop/crystal")
    assert crystal_list.status_code == 200 and crystal_order_no in crystal_list.text
    assert "data-selected-amount-total" not in crystal_list.text and "&#21333;&#20215;" not in crystal_list.text
    polishing_unlock = client.post(
        "/workshop/polishing/unlock",
        data={"csrf": csrf(home.text), "password": "polishing-pass-123"},
        follow_redirects=False,
    )
    assert polishing_unlock.status_code == 303 and polishing_unlock.headers["location"] == "/workshop/polishing"
    polishing = client.get("/workshop/polishing")
    assert polishing.status_code == 200 and "touch-piecework-panel" in polishing.text
    assert 'name="unit_price"' in polishing.text and 'data-workshop-employees' in polishing.text
    assert "\u725f\u6c5f" in polishing.text and "\u6bdb\u536b\u5175" in polishing.text
    assert 'employee-button active' not in polishing.text
    assert "\u6253\u6837\u8d39" in polishing.text and 'name="mold_fee"' in polishing.text
    polishing_report = client.post(
        "/workshop/polishing",
        data={"csrf": csrf(polishing.text), "employee_name": ["\u725f\u6c5f"], "note_text": ["\u65e0"], "order_no": [polishing_order_no], "quantity": ["18"], "unit_price": ["0.2"], "mold_fee": ["10"]},
        follow_redirects=False,
    )
    assert polishing_report.status_code == 303
    polishing_records = repo.order_workshop_records(polishing_order_id)
    assert len(polishing_records) == 1
    assert {row["operator_name"] for row in polishing_records} == {"\u725f\u6c5f"}
    for row in polishing_records:
        assert row["department_name"] == "\u629b\u5149"
        assert row["quantity"] == 18
        assert abs(row["unit_price"] - 0.2) < 1e-9
        assert abs(row["mold_fee"] - 10) < 1e-9
        assert abs(row["amount"] - 13.6) < 1e-9
    polishing_list = client.get("/workshop/polishing")
    assert polishing_list.status_code == 200 and polishing_order_no in polishing_list.text
    assert "data-selected-amount-total" in polishing_list.text and 'data-amount="13.60"' in polishing_list.text
    assert 'data-selection-amount-total-all="13.60"' in polishing_list.text
    painting_unlock = client.post(
        "/workshop/painting/unlock",
        data={"csrf": csrf(home.text), "password": "painting-pass-123"},
        follow_redirects=False,
    )
    assert painting_unlock.status_code == 303 and painting_unlock.headers["location"] == "/workshop/painting"
    painting = client.get("/workshop/painting")
    assert painting.status_code == 200 and "touch-piecework-panel" in painting.text
    assert "\u989c\u8272\u6570\u91cf" in painting.text and 'data-default-value="1"' in painting.text
    assert 'employee-button active' not in painting.text
    assert "\u5218\u8fdb" in painting.text and "\u5f90\u53cb\u4e3d" in painting.text
    painting_report = client.post(
        "/workshop/painting",
        data={"csrf": csrf(painting.text), "employee_name": ["\u5218\u8fdb,\u9ec4\u4e09\u679a"], "order_no": [painting_order_no], "quantity": ["100"], "unit_price": ["0.15"], "mold_fee": ["3"]},
        follow_redirects=False,
    )
    assert painting_report.status_code == 303
    painting_records = repo.order_workshop_records(painting_order_id)
    assert len(painting_records) == 2
    for row in painting_records:
        assert row["department_name"] == "\u4e0a\u8272"
        assert row["quantity"] == 50
        assert abs(row["unit_price"] - 0.15) < 1e-9
        assert abs(row["mold_fee"] - 3) < 1e-9
        assert abs(row["amount"] - 22.5) < 1e-9
    painting_list = client.get("/workshop/painting")
    assert painting_list.status_code == 200 and 'data-selection-amount-total-all="45.00"' in painting_list.text
    diecast_unlock = client.post(
        "/workshop/diecast/unlock",
        data={"csrf": csrf(home.text), "password": "diecast-pass-123"},
        follow_redirects=False,
    )
    assert diecast_unlock.status_code == 303 and diecast_unlock.headers["location"] == "/workshop/diecast"
    diecast = client.get("/workshop/diecast")
    assert diecast.status_code == 200 and "touch-piecework-panel" in diecast.text
    assert "\u519c\u5982\u5e72" in diecast.text and "\u519c\u5929\u4f69" in diecast.text
    assert "\u88c5\u6a21\u8d39" in diecast.text and 'list="mold-fee-options-diecast"' in diecast.text and '<option value="8"></option>' in diecast.text
    assert 'employee-button active' not in diecast.text
    diecast_report = client.post(
        "/workshop/diecast",
        data={"csrf": csrf(diecast.text), "employee_name": ["\u519c\u5982\u5e72,\u674e\u56fd\u5bcc"], "order_no": [diecast_order_no], "quantity": ["80"], "unit_price": ["0.25"], "mold_fee": ["9"]},
        follow_redirects=False,
    )
    assert diecast_report.status_code == 303
    diecast_records = repo.order_workshop_records(diecast_order_id)
    assert len(diecast_records) == 2
    for row in diecast_records:
        assert row["department_name"] == "\u538b\u94f8"
        assert row["quantity"] == 40
        assert abs(row["unit_price"] - 0.25) < 1e-9
        assert abs(row["mold_fee"] - 4.5) < 1e-9
        assert abs(row["amount"] - 14.5) < 1e-9
    diecast_list = client.get("/workshop/diecast")
    assert diecast_list.status_code == 200 and 'data-selection-amount-total-all="29.00"' in diecast_list.text
    packaging_unlock = client.post(
        "/workshop/packaging/unlock",
        data={"csrf": csrf(home.text), "password": "packaging-pass-123"},
        follow_redirects=False,
    )
    assert packaging_unlock.status_code == 303 and packaging_unlock.headers["location"] == "/workshop/packaging"
    packaging = client.get("/workshop/packaging")
    assert packaging.status_code == 200 and "touch-piecework-panel" in packaging.text
    assert 'name="mold_fee"' not in packaging.text and 'data-batch-workshop-price' not in packaging.text
    packaging_employees = ["\u6d82\u5c0f\u82f1", "\u5f90\u5f69\u8fde", "\u5468\u7f8e\u8bc6", "\u9648\u5c0f\u971e", "\u738b\u5bb6\u4e3d", "\u6768\u660e\u4ed9", "\u5f20\u96ea\u6797", "\u738b\u6587\u5bb9", "\u66fe\u51e4\u5a25", "\u66fe\u8fde\u543e", "\u8d56\u706b\u80dc"]
    for employee in packaging_employees:
        assert f'data-employee-value="{employee}"' in packaging.text
        assert f'<option value="{employee}"' in packaging.text
    packaging_report = client.post(
        "/workshop/packaging",
        data={"csrf": csrf(packaging.text), "employee_name": ["\u6d82\u5c0f\u82f1,\u5f90\u5f69\u8fde"], "order_no": [packaging_order_no], "quantity": ["100"], "unit_price": ["0.5"]},
        follow_redirects=False,
    )
    assert packaging_report.status_code == 303
    packaging_records = repo.order_workshop_records(packaging_order_id)
    assert len(packaging_records) == 2
    assert {row["operator_name"] for row in packaging_records} == {"\u6d82\u5c0f\u82f1", "\u5f90\u5f69\u8fde"}
    for row in packaging_records:
        assert row["department_name"] == "\u5305\u88c5"
        assert row["quantity"] == 50
        assert abs(row["unit_price"] - 0.5) < 1e-9
        assert abs(row["mold_fee"] - 0) < 1e-9
        assert abs(row["amount"] - 25) < 1e-9
    packaging_list = client.get("/workshop/packaging")
    assert packaging_list.status_code == 200 and packaging_order_no in packaging_list.text
    assert 'data-amount="25.00"' in packaging_list.text and 'data-selection-amount-total-all="50.00"' in packaging_list.text
    assert "&#35013;&#27169;&#36153;" not in packaging_list.text
    packaging_export = client.post(
        "/workshop/packaging/export",
        data={"csrf": csrf(packaging_list.text), "selected_ids": [str(row["id"]) for row in packaging_records]},
    )
    assert packaging_export.status_code == 200
    packaging_workbook_path = root / "packaging-export.xlsx"
    packaging_workbook_path.write_bytes(packaging_export.content)
    packaging_sheet = load_workbook(packaging_workbook_path).active
    assert packaging_sheet.cell(row=1, column=6).value == "\u53c2\u8003\u6570\u91cf"
    assert packaging_sheet.cell(row=1, column=8).value == "\u91d1\u989d"
    assert packaging_sheet.cell(row=1, column=9).value == "\u62a5\u5230\u65f6\u95f4"
    assert packaging_sheet.cell(row=1, column=10).value == "\u4ea7\u54c1\u7f29\u7565\u56fe"
    assert packaging_sheet.cell(row=2, column=1).value == packaging_order_no
    list_page = client.get("/workshop/mold")
    assert 'type="date"' in list_page.text
    assert "operator_name" not in list_page.text and "&#25805;&#20316;&#20154;" not in list_page.text
    assert "&#20986;&#36135;&#29366;&#24577;" not in list_page.text
    assert "/workshop/mold/ship" not in list_page.text
    assert 'data-delete-url="/workshop/mold/records/' not in list_page.text
    assert "&#26448;&#36136;" in list_page.text and "&#23610;&#23544;" in list_page.text and "&#35268;&#26684;" in list_page.text and "&#35746;&#21333;&#31867;&#21035;" in list_page.text
    assert "\u950c" in list_page.text and "50MM" in list_page.text and "2D" in list_page.text and "&#27491;&#24120;" in list_page.text
    assert ">&#20135;&#21697;<" not in list_page.text and ">&#23458;&#25143;<" not in list_page.text and ">&#37096;&#38376;<" not in list_page.text
    assert 'data-request-edit-url="/workshop/mold/records/' not in list_page.text
    assert 'data-request-edit-mode="prompt"' not in list_page.text
    assert 'data-workshop-quantity-url="/workshop/mold/records/' in list_page.text
    assert 'data-workshop-quantity="2"' in list_page.text
    assert 'data-select-all' in list_page.text and 'data-requires-selection' in list_page.text
    wide_date_page = client.get("/workshop/mold?reported_from=1900-01-01&reported_to=2999-12-31")
    assert wide_date_page.status_code == 200 and order_no in wide_date_page.text
    narrow_date_page = client.get("/workshop/mold?reported_from=1900-01-01&reported_to=1900-01-01")
    assert narrow_date_page.status_code == 200 and order_no not in narrow_date_page.text
    history = client.get(f"/workshop/mold/history?order_no={order_no}")
    assert history.status_code == 200
    assert history.json()["record"]["quantity"] == 1
    assert abs(history.json()["record"]["unit_price"] - 10.5) < 1e-9
    assert history.json()["record"]["material"] == "锌"
    assert history.json()["record"]["size_text"] == "\u9ad81\u5bbd20\u539a5"
    assert history.json()["record"]["spec"] == "2D+\u80cc\u5b57"
    quantity_request = client.post(
        f"/workshop/mold/records/{records[0]['id']}/quantity/request",
        data={"csrf": csrf(list_page.text), "quantity": "5", "unit_price": "12.5", "reason": "漏扫数量"},
        follow_redirects=False,
    )
    assert quantity_request.status_code == 303 and quantity_request.headers["location"] == "/messages"
    assert repo.order_workshop_records(order_id)[0]["quantity"] == 2
    assert abs(repo.order_workshop_records(order_id)[0]["unit_price"] - 10.5) < 1e-9
    workshop_messages = client.get("/messages")
    assert workshop_messages.status_code == 200 and "刻模数量修改" in workshop_messages.text and "数量从2修改为5" in workshop_messages.text
    duplicate_report = client.post(
        "/workshop/mold",
        data={"csrf": csrf(client.get("/workshop/mold").text), "order_no": [order_no], "material": ["锌"], "size_text": ["50MM"], "spec": ["2D"], "quantity": ["3"], "unit_price": ["20"], "record_type": ["normal"]},
        follow_redirects=False,
    )
    assert duplicate_report.status_code == 422
    records = repo.order_workshop_records(order_id)
    assert len(records) == 1
    rework_report = client.post(
        "/workshop/mold",
        data={"csrf": csrf(client.get("/workshop/mold").text), "order_no": [order_no], "material": ["铜"], "size_text": ["52MM"], "spec": ["3D"], "quantity": ["3"], "unit_price": ["20"], "record_type": ["rework"]},
        follow_redirects=False,
    )
    assert rework_report.status_code == 303
    records = repo.order_workshop_records(order_id)
    assert len(records) == 2
    assert records[1]["quantity"] == 3
    assert records[1]["material"] == "铜"
    assert records[1]["record_type"] == "rework"
    assert abs(records[1]["unit_price"] - 20) < 1e-9
    history = client.get(f"/workshop/mold/history?order_no={order_no}")
    assert history.json()["record"]["quantity"] == 1
    assert abs(history.json()["record"]["unit_price"] - 20) < 1e-9
    assert history.json()["record"]["record_type"] == "rework"
    delete = client.post(
        f"/workshop/mold/records/{records[1]['id']}/delete",
        data={"csrf": csrf(client.get("/workshop/mold").text)},
        follow_redirects=False,
    )
    assert delete.status_code == 403
    records = repo.order_workshop_records(order_id)
    assert len(records) == 2
    client.post("/logout", data={"csrf": csrf(client.get("/workshop/mold").text)})
    admin_login_page = client.get("/login")
    admin_login = client.post(
        "/login",
        data={"csrf": csrf(admin_login_page.text), "username": "admin", "password": "admin-pass-123"},
        follow_redirects=False,
    )
    assert admin_login.status_code == 303
    admin_press_detail = client.get(f"/orders/{press_order_id}")
    assert admin_press_detail.status_code == 200
    assert "\u51b2\u538b&#65288;\u5f90\u5c71\u7acb&#65289;" in admin_press_detail.text
    assert "\u51b2\u538b&#65288;\u5218\u9053\u6797&#65289;" in admin_press_detail.text
    admin_mold_detail = client.get(f"/orders/{order_id}")
    assert admin_mold_detail.status_code == 200
    assert "\u523b\u6a21&#65288;" not in admin_mold_detail.text
    admin_mold_list = client.get("/workshop/mold")
    assert 'data-delete-url="/workshop/mold/records/' in admin_mold_list.text
    assert 'data-csrf="' in admin_mold_list.text
    admin_delete = client.post(
        f"/workshop/mold/records/{records[1]['id']}/delete",
        data={"csrf": csrf(admin_mold_list.text)},
        follow_redirects=False,
    )
    assert admin_delete.status_code == 303
    records = repo.order_workshop_records(order_id)
    assert len(records) == 1

    client.post("/logout", data={"csrf": csrf(admin_mold_list.text)})
    workshop_login_page = client.get("/login")
    workshop_login = client.post(
        "/login",
        data={"csrf": csrf(workshop_login_page.text), "username": "workshop", "password": "workshop-pass-123"},
        follow_redirects=False,
    )
    assert workshop_login.status_code == 303
    detail = client.get(f"/orders/{order_id}")
    assert detail.status_code == 200
    assert_workshop_detail_pdf_only(detail.text, order_id, "10.5000")
    cutter_detail = client.get(f"/orders/{cutter_order_id}")
    assert cutter_detail.status_code == 200
    assert_workshop_detail_pdf_only(cutter_detail.text, cutter_order_id, "8.8000")
    workshop_user = repo.get_user(2)
    assert workshop_user is not None
    for index in range(45):
        _, bulk_order_no = repo.create_order(payload(f"TWD1-260722{index + 200:03d}"))
        repo.create_workshop_records(
            "mold",
            "\u523b\u6a21",
            [{"order_no": bulk_order_no, "material": "锌", "size_text": "45MM", "spec": "2D", "quantity": 1, "unit_price": 6.6, "record_type": "normal"}],
            workshop_user,
        )
    bulk_home = client.get("/workshop")
    bulk_unlock = client.post(
        "/workshop/mold/unlock",
        data={"csrf": csrf(bulk_home.text), "password": "mold-pass-123"},
        follow_redirects=False,
    )
    assert bulk_unlock.status_code == 303
    bulk_page = client.get("/workshop/mold?q=TWD1-260722")
    assert bulk_page.status_code == 200 and 'data-selection-total="45"' in bulk_page.text
    bulk_export = client.post(
        "/workshop/mold/export",
        data={"csrf": csrf(bulk_page.text), "select_scope": "all_matching", "q": "TWD1-260722"},
    )
    assert bulk_export.status_code == 200
    bulk_workbook_path = root / "mold-bulk-export.xlsx"
    bulk_workbook_path.write_bytes(bulk_export.content)
    bulk_sheet = load_workbook(bulk_workbook_path).active
    exported_order_nos = {bulk_sheet.cell(row=row_index, column=1).value for row_index in range(2, bulk_sheet.max_row + 1)}
    assert len(exported_order_nos) == 45
    assert "TWD1-260722200" in exported_order_nos and "TWD1-260722244" in exported_order_nos
    cutter_export_home = client.get("/workshop")
    cutter_export_unlock = client.post(
        "/workshop/cutter/unlock",
        data={"csrf": csrf(cutter_export_home.text), "password": "cutter-pass-123"},
        follow_redirects=False,
    )
    assert cutter_export_unlock.status_code == 303
    cutter_export = client.post(
        "/workshop/cutter/export",
        data={"csrf": csrf(client.get("/workshop/cutter").text), "selected_ids": [str(cutter_records[0]["id"])]},
    )
    assert cutter_export.status_code == 200
    assert "spreadsheetml.sheet" in cutter_export.headers["content-type"]
    workbook_path = root / "cutter-export.xlsx"
    workbook_path.write_bytes(cutter_export.content)
    sheet = load_workbook(workbook_path).active
    headers = [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]
    assert sheet.max_column == 8
    assert not any("\u51fa\u8d27" in str(value) or "鍑鸿揣" in str(value) for value in headers)
    assert sheet.cell(row=2, column=1).value == cutter_order_no
    assert sheet.cell(row=2, column=3).value == "\u7279\u6b8a"
    assert sheet.cell(row=2, column=2).value == "35MM"
    assert sheet.cell(row=2, column=5).value == 8.8

    page = client.get("/")
    client.post("/logout", data={"csrf": csrf(page.text)})
    login_page = client.get("/login")
    login = client.post(
        "/login",
        data={"csrf": csrf(login_page.text), "username": "admin", "password": "admin-pass-123"},
        follow_redirects=False,
    )
    assert login.status_code == 303

    admin_messages = client.get("/messages")
    assert admin_messages.status_code == 200 and "刻模数量修改" in admin_messages.text
    review = client.post(
        "/messages/1/review",
        data={"csrf": csrf(admin_messages.text), "decision": "approve", "review_note": "同意"},
        follow_redirects=False,
    )
    assert review.status_code == 303
    assert repo.order_workshop_records(order_id)[0]["quantity"] == 5
    assert abs(repo.order_workshop_records(order_id)[0]["unit_price"] - 12.5) < 1e-9

    admin_detail = client.get(f"/orders/{order_id}")
    assert admin_detail.status_code == 200
    assert "workflow-line" in admin_detail.text
    assert "current" in admin_detail.text
    assert "刻模" in admin_detail.text
    assert "冲压" not in admin_detail.text
    assert "12.5000" in admin_detail.text
    assert ">5<" in admin_detail.text

print(f"workshop smoke ok: {root}")
